"""Reading a catalogue out of a spreadsheet.

Every institution that has ever catalogued anything arrives with an Excel file.
It was made by somebody who left, its column headings are abbreviations, its
dates are half text and half real dates, and it is the only record of four
thousand objects. Refusing it is not an option, and neither is trusting it.

So the import is three separate steps, and the middle one is the point:

1. **Analyse.** Read the file. Report what columns it has, what is in them, and
   a *guess* at which platform field each one fills.
2. **Verify.** A person confirms or corrects that mapping, column by column.
   The platform never decides this silently — a column headed "Date" could be
   the acquisition date, the date of manufacture, or the date somebody typed
   the row, and only the cataloguer knows which.
3. **Preview, then commit.** With the mapping settled, every row is validated
   and the result reported *before* anything is written. A file that would
   create eight hundred objects and fail on three says so, and says which
   three, while the database is still untouched.

The field descriptions come from :mod:`app.services.forms` — the same layout
the cataloguing card is drawn from. A column can therefore fill exactly what a
person could have typed, and there is no second list of importable fields to
drift out of step with the form.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from app.services import forms

#: How many rows of the file to keep as examples for the mapping screen. Enough
#: to see what a column really holds; few enough to send in a page.
SAMPLE_ROWS = 5

#: Refuse a file larger than this. A catalogue is text; anything this size is
#: either embedded images or a mistake.
MAX_BYTES = 25 * 1024 * 1024

#: Refuse a sheet with more rows than this in one go.
MAX_ROWS = 50_000

#: Refuse a sheet wider than this. Beyond it, the column-mapping screen is not
#: usable and the file is almost certainly not a catalogue.
MAX_COLUMNS = 200


class SpreadsheetError(Exception):
    """Something about the file makes it unreadable. The message is for a person."""


# --------------------------------------------------------------------------
# Reading
# --------------------------------------------------------------------------
@dataclass
class Sheet:
    """One table read out of a file."""

    name: str
    #: Column headings, in file order, with blanks named for their position.
    columns: list[str]
    rows: list[dict[str, Any]]
    #: Which row the headings came from, 1-based. Worth carrying because it may
    #: have been *worked out* rather than given, and the screen that shows the
    #: mapping has to say which row it read and let somebody move it.
    header_row: int = 1

    @property
    def sample(self) -> list[dict[str, Any]]:
        return self.rows[:SAMPLE_ROWS]


def _clean_header(value: Any, index: int) -> str:
    """A usable name for a column, even when the file gives none."""
    text = "" if value is None else str(value).strip()
    # Excel exports carry non-breaking spaces and stray newlines in headings.
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text or f"Column {index + 1}"


def _unique(names: list[str]) -> list[str]:
    """Two columns headed "Notes" are common, and would otherwise collide."""
    seen: dict[str, int] = {}
    result: list[str] = []
    for name in names:
        count = seen.get(name, 0)
        seen[name] = count + 1
        result.append(name if count == 0 else f"{name} ({count + 1})")
    return result


def read_csv(data: bytes, *, header_row: int | None = None) -> Sheet:
    """Read a delimited text file, guessing its delimiter and encoding."""
    text = _decode(data)

    # Sniff on a slice: a whole file confuses the sniffer more than it helps.
    sample = text[:8192]
    try:
        dialect: Any = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    all_rows = list(reader)
    if header_row is None:
        header_row = find_header_row([list(row) for row in all_rows])
    if len(all_rows) < header_row:
        raise SpreadsheetError(
            f"The file has {len(all_rows)} rows, so there is no row {header_row} "
            f"to read column headings from."
        )

    header = all_rows[header_row - 1]
    columns = _unique([_clean_header(value, index) for index, value in enumerate(header)])
    _check_shape(columns, len(all_rows) - header_row)

    rows = []
    for raw in all_rows[header_row:]:
        if not any(str(cell).strip() for cell in raw):
            continue  # A blank separator row is not a record.
        rows.append(
            {name: _normalise(raw[i] if i < len(raw) else None) for i, name in enumerate(columns)}
        )

    return Sheet(name="CSV", columns=columns, rows=rows, header_row=header_row)


def _decode(data: bytes) -> str:
    """Excel writes UTF-8 with a BOM, or Windows-1252, and never says which."""
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise SpreadsheetError(
        "The file's text could not be decoded. Save it again as CSV UTF-8, or "
        "as .xlsx, and try once more."
    )


def read_xlsx(
    data: bytes, *, sheet_name: str | None = None, header_row: int | None = None
) -> Sheet:
    """Read one worksheet of an Excel file."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise SpreadsheetError("Excel support is not installed on this server.") from exc

    try:
        # read_only keeps memory flat on a large file; data_only takes the last
        # calculated value of a formula rather than the formula's text, which is
        # what a person reading the file in Excel sees.
        book = openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        raise SpreadsheetError(
            "That file could not be opened as an Excel workbook. If it is an "
            "old .xls, open it in Excel and save it as .xlsx."
        ) from exc

    try:
        if sheet_name is not None:
            if sheet_name not in book.sheetnames:
                raise SpreadsheetError(
                    f"The workbook has no sheet named {sheet_name!r}. It has: "
                    f"{', '.join(book.sheetnames)}."
                )
            worksheet = book[sheet_name]
        else:
            worksheet = book[book.sheetnames[0]]

        iterator = worksheet.iter_rows(values_only=True)

        if header_row is None:
            # Peek at the top of the sheet, work out where the headings are,
            # then start again. Cheap: a handful of rows, once.
            peek = []
            for _ in range(HEADER_SEARCH_ROWS):
                row = next(iterator, None)
                if row is None:
                    break
                peek.append(list(row))
            header_row = find_header_row(peek) if peek else 1
            iterator = worksheet.iter_rows(values_only=True)

        header: tuple[Any, ...] | None = None
        for _ in range(header_row):
            header = next(iterator, None)
        if header is None:
            raise SpreadsheetError(
                f"The sheet has fewer than {header_row} rows, so there is no row "
                f"{header_row} to read column headings from."
            )

        columns = _unique([_clean_header(value, index) for index, value in enumerate(header)])

        rows: list[dict[str, Any]] = []
        for raw in iterator:
            if raw is None or not any(cell is not None and str(cell).strip() for cell in raw):
                continue
            rows.append(
                {
                    name: _normalise(raw[i] if i < len(raw) else None)
                    for i, name in enumerate(columns)
                }
            )
            if len(rows) > MAX_ROWS:
                raise SpreadsheetError(
                    f"The sheet has more than {MAX_ROWS:,} rows. Split it and "
                    f"import the parts separately."
                )

        _check_shape(columns, len(rows))
        return Sheet(name=worksheet.title, columns=columns, rows=rows, header_row=header_row)
    finally:
        book.close()


def _check_shape(columns: list[str], row_count: int) -> None:
    if len(columns) > MAX_COLUMNS:
        raise SpreadsheetError(
            f"The sheet has {len(columns)} columns, more than the {MAX_COLUMNS} "
            f"this can map. Delete the columns you do not need and try again."
        )
    if row_count == 0:
        raise SpreadsheetError(
            "There are no data rows below the heading row. Check that the "
            "heading row number is right."
        )


def _normalise(value: Any) -> Any:
    """Make a cell into something JSON can carry, without losing meaning."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return (
            value.date().isoformat() if value.time() == datetime.min.time() else value.isoformat()
        )
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        text = value.replace("\xa0", " ").strip()
        return text or None
    return value


def sheet_names(data: bytes) -> list[str]:
    """The sheets in a workbook, so the caller can choose one."""
    import openpyxl

    try:
        book = openpyxl.load_workbook(io.BytesIO(data), read_only=True, keep_links=False)
    except Exception as exc:
        raise SpreadsheetError("That file could not be opened as an Excel workbook.") from exc
    try:
        return list(book.sheetnames)
    finally:
        book.close()


# --------------------------------------------------------------------------
# Finding the headings
# --------------------------------------------------------------------------
#: How far down to look. A title, a blank line, a subtitle and a merged banner
#: is four; beyond about eight the file is something other than a table and
#: guessing further would be guessing.
HEADER_SEARCH_ROWS = 8

#: How much better a later row must score before the headings are taken from
#: it rather than from row 1. Deliberately wide.
MOVE_MARGIN = 0.75


def _header_score(row: list[Any], below: list[list[Any]]) -> float:
    """How much this row looks like a row of column headings.

    A heading row is short text, most cells filled, few numbers, no repeats -
    and, crucially, *unlike the rows under it*. A row of "Stone", "Stone",
    "Copper Alloy" is data no matter how word-like it is; a row of "Material",
    "Height (cm)", "Weight (gr)" is not, because nothing repeats and the rows
    below it are full of numbers where it has none.
    """
    values = ["" if cell is None else str(cell).strip() for cell in row]
    filled = [value for value in values if value]
    if len(filled) < 2:
        return 0.0

    score = 0.0
    # Most of the row has something in it.
    score += 2.0 * (len(filled) / max(len(values), 1))

    # Headings are words. A row that is mostly numbers is data.
    numeric = sum(1 for value in filled if _looks_numeric(value))
    score += 2.0 * (1 - numeric / len(filled))

    # Headings are distinct. Data repeats.
    score += 1.5 * (len({value.lower() for value in filled}) / len(filled))

    # Headings are short.
    average = sum(len(value) for value in filled) / len(filled)
    if average <= 40:
        score += 1.0
    if average > 80:
        score -= 1.0

    # And the strongest signal of all: the rows underneath contain numbers
    # where this row does not.
    for other in below[:5]:
        others = ["" if cell is None else str(cell).strip() for cell in other]
        other_filled = [value for value in others if value]
        if not other_filled:
            continue
        other_numeric = sum(1 for value in other_filled if _looks_numeric(value))
        if other_numeric / len(other_filled) > numeric / len(filled):
            score += 0.4

    return score


def _looks_numeric(value: str) -> bool:
    try:
        float(value.replace(",", "").replace(" ", ""))
    except ValueError:
        return False
    return True


def find_header_row(rows: list[list[Any]]) -> int:
    """Which row holds the column headings, 1-based.

    Assuming row 1 is right most of the time and catastrophic the rest of it: a
    file with a title above the table gets columns called "Column 6", the real
    headings become the first row of data, and every single row then fails for
    reasons that are actually the heading text - "'Material' is not one of the
    values this field accepts". Nothing about that says "the headings are one
    row further down", which is the only thing anybody needed to be told.

    So: score the first few rows and take the best. It is a suggestion, shown
    on the mapping screen and changeable there - the reader never insists.
    """
    if not rows:
        return 1

    scores: list[float] = []
    for index in range(min(HEADER_SEARCH_ROWS, len(rows))):
        below = rows[index + 1 :]
        # A heading row with nothing under it is not a heading row. Without
        # this a two-line file - one heading, one record - can be read as one
        # heading on line 2 and no records at all.
        if not any(any(str(cell).strip() for cell in row if cell is not None) for row in below):
            scores.append(-1.0)
            continue
        scores.append(_header_score(rows[index], below))

    # Row 1 unless something further down is clearly better. Most files start
    # at the top, the scoring is a heuristic, and a heuristic that overrules
    # the ordinary case on a narrow margin is worse than no heuristic: it
    # would move the headings of a perfectly good file for no reason anybody
    # could see.
    best_index = 0
    for index in range(1, len(scores)):
        if scores[index] > scores[best_index] + MOVE_MARGIN:
            best_index = index
    return best_index + 1


def read(
    data: bytes, *, filename: str, sheet_name: str | None = None, header_row: int | None = None
) -> Sheet:
    """Read whichever kind of file this is."""
    if len(data) > MAX_BYTES:
        raise SpreadsheetError(
            f"The file is larger than {MAX_BYTES // (1024 * 1024)} MB. A catalogue "
            f"of text should be far smaller; check it does not contain images."
        )
    if header_row is not None and header_row < 1:
        raise SpreadsheetError("The heading row must be row 1 or later.")

    lowered = filename.lower()
    if lowered.endswith((".xlsx", ".xlsm")):
        # `header_row=None` means "work it out", which needs the raw grid, so
        # the detection happens inside each reader rather than here.
        return read_xlsx(data, sheet_name=sheet_name, header_row=header_row)
    if lowered.endswith((".csv", ".tsv", ".txt")):
        return read_csv(data, header_row=header_row)
    if lowered.endswith(".xls"):
        raise SpreadsheetError(
            "That is the old Excel format (.xls). Open it in Excel or LibreOffice "
            "and save it as .xlsx, then try again."
        )
    raise SpreadsheetError(f"{filename!r} is not a spreadsheet this can read. Send .xlsx or .csv.")


# --------------------------------------------------------------------------
# Guessing what a column is
# --------------------------------------------------------------------------
def _slug(value: str) -> str:
    """Strip a heading down to something comparable across spellings."""
    text = unicodedata.normalize("NFKD", value)
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


#: Headings seen in real museum spreadsheets, mapped to the field they mean.
#: This is a convenience, not a decision: every guess is shown to a person to
#: confirm before a single row is written.
SYNONYMS: dict[str, dict[str, str]] = {
    "museum_object": {
        "accessionno": "accession_number",
        "accessionnumber": "accession_number",
        "accno": "accession_number",
        "acc": "accession_number",
        "registrationnumber": "accession_number",
        "regno": "accession_number",
        "inventorynumber": "accession_number",
        "invno": "accession_number",
        "objectnumber": "accession_number",
        "formernumber": "former_number",
        "formerno": "former_number",
        "oldnumber": "former_number",
        "oldno": "former_number",
        "previousnumber": "former_number",
        "title": "title",
        "name": "title",
        "objectname": "title",
        "object": "title",
        "objecttype": "object_type",
        "type": "object_type",
        "classification": "object_type",
        "description": "description",
        "notes": "description",
        "remarks": "description",
        "culture": "culture",
        "cultureperiod": "culture",
        "period": "period_id",
        "date": "date_note",
        "dating": "date_note",
        "daterange": "date_note",
        "datefrom": "date_from",
        "earliestdate": "date_from",
        "dateto": "date_to",
        "latestdate": "date_to",
        "material": "materials",
        "materials": "materials",
        "medium": "materials",
        "technique": "techniques",
        "techniques": "techniques",
        "maker": "maker",
        "artist": "maker",
        "manufacturer": "maker",
        "inscription": "inscription",
        "inscriptions": "inscription",
        "marks": "marks",
        "height": "height_mm",
        "heightmm": "height_mm",
        "h": "height_mm",
        "width": "width_mm",
        "widthmm": "width_mm",
        "w": "width_mm",
        "depth": "depth_mm",
        "depthmm": "depth_mm",
        "d": "depth_mm",
        "diameter": "diameter_mm",
        "diam": "diameter_mm",
        "thickness": "thickness_mm",
        "weight": "weight_g",
        "weightg": "weight_g",
        "condition": "condition",
        "conditionnote": "condition_note",
        "conservationstatus": "conservation_status",
        "acquisitionmethod": "acquisition_method",
        "howacquired": "acquisition_method",
        "acquisitiondate": "acquisition_date",
        "dateacquired": "acquisition_date",
        "acquiredfrom": "acquired_from",
        "source": "acquired_from",
        "donor": "acquired_from",
        "provenance": "provenance",
        "findspot": "find_spot",
        "location": "storage_location_id",
        "currentlocation": "storage_location_id",
        "storagelocation": "storage_location_id",
        "status": "status",
        "collection": "collection_id",
        "category": "category_id",
    },
    # ----------------------------------------------------------------------
    # The excavation record types.
    #
    # Written from the column headings that actually appear on a site's
    # spreadsheets rather than from the field names, which is the whole point:
    # a register that says "SU" or "Locus" or "Ctx" should not have to be
    # rewritten before it can be read.
    # ----------------------------------------------------------------------
    "site": {
        "sitecode": "code",
        "code": "code",
        "siteno": "code",
        "sitenumber": "code",
        "sitename": "name",
        "name": "name",
        "alsoknownas": "alternative_names",
        "alternativenames": "alternative_names",
        "othernames": "alternative_names",
        "sitetype": "site_type",
        "type": "site_type",
        "description": "description",
        "lat": "latitude",
        "latitude": "latitude",
        "y": "latitude",
        "lon": "longitude",
        "lng": "longitude",
        "long": "longitude",
        "longitude": "longitude",
        "x": "longitude",
        "elevation": "elevation",
        "altitude": "elevation",
        "masl": "elevation",
        "accuracy": "location_accuracy_m",
        "country": "country",
        "region": "region",
        "governorate": "region",
        "province": "region",
        "district": "district",
        "municipality": "municipality",
        "address": "address",
        "registerno": "national_register_id",
        "nationalregisterno": "national_register_id",
        "period": "period_id",
        "periodtext": "period_text",
        "dating": "period_text",
        "datingmethod": "dating_method",
        "datefrom": "date_from",
        "dateto": "date_to",
        "condition": "condition",
        "protection": "protection_status",
        "protectionstatus": "protection_status",
        "threats": "threats",
        "landuse": "land_use",
        "landowner": "landowner",
        "owner": "landowner",
        "discovered": "discovery_date",
        "discoverydate": "discovery_date",
        "discoveredby": "discovered_by",
        "excavationstart": "excavation_start",
        "excavationend": "excavation_end",
        "references": "references",
        "bibliography": "references",
        "notes": "notes",
        "remarks": "notes",
        "keywords": "keywords",
    },
    "excavation_context": {
        "context": "context_number",
        "contextno": "context_number",
        "contextnumber": "context_number",
        "ctx": "context_number",
        "locus": "context_number",
        "locusno": "context_number",
        "su": "context_number",
        "unit": "context_number",
        "contexttype": "context_type",
        "type": "context_type",
        "locustype": "context_type",
        "stratigraphicunit": "stratigraphic_unit",
        "stratum": "stratigraphic_unit",
        "phase": "phase",
        "trench": "trench",
        "area": "area",
        "square": "square",
        "grid": "square",
        "gridsquare": "square",
        "description": "description",
        "interpretation": "interpretation",
        "munsell": "munsell_color",
        "munsellcolour": "munsell_color",
        "munsellcolor": "munsell_color",
        "colour": "munsell_color",
        "color": "munsell_color",
        "composition": "composition",
        "soil": "composition",
        "compaction": "compaction",
        "inclusions": "inclusions",
        "length": "length_cm",
        "width": "width_cm",
        "depth": "depth_cm",
        "thickness": "thickness_cm",
        "topelevation": "top_elevation",
        "toplevel": "top_elevation",
        "bottomelevation": "bottom_elevation",
        "bottomlevel": "bottom_elevation",
        "lat": "latitude",
        "latitude": "latitude",
        "lon": "longitude",
        "longitude": "longitude",
        "excavatedby": "excavated_by",
        "dugby": "excavated_by",
        "excavationdate": "excavation_date",
        "date": "excavation_date",
        "recordedby": "recorded_by",
        "recorder": "recorded_by",
        "samples": "samples_taken",
        "period": "period_id",
        "datingevidence": "dating_evidence",
        "datefrom": "date_from",
        "dateto": "date_to",
        "notes": "notes",
        "remarks": "notes",
    },
    "artifact": {
        "invno": "inventory_number",
        "inventoryno": "inventory_number",
        "inventorynumber": "inventory_number",
        "findno": "inventory_number",
        "findnumber": "inventory_number",
        "objectno": "inventory_number",
        "regno": "inventory_number",
        "fieldno": "field_number",
        "fieldnumber": "field_number",
        "basketno": "field_number",
        "basket": "field_number",
        "bagno": "field_number",
        "accessionno": "accession_number",
        "museumno": "accession_number",
        "name": "name",
        "objectname": "name",
        "find": "name",
        "objecttype": "object_type",
        "type": "object_type",
        "classification": "object_type",
        "ware": "typology",
        "typology": "typology",
        "form": "typology",
        "category": "category_id",
        "description": "description",
        "quantity": "quantity",
        "count": "quantity",
        "sherdcount": "quantity",
        "no": "quantity",
        "fragment": "is_fragment",
        "context": "context_id",
        "contextno": "context_id",
        "locus": "context_id",
        "su": "context_id",
        "stratigraphicunit": "stratigraphic_unit",
        "stratum": "stratigraphic_unit",
        "trench": "trench",
        "square": "square",
        "grid": "square",
        "depth": "depth_cm",
        "elevation": "elevation",
        "level": "elevation",
        "lat": "latitude",
        "latitude": "latitude",
        "lon": "longitude",
        "longitude": "longitude",
        "finddate": "find_date",
        "datefound": "find_date",
        "foundby": "found_by",
        "excavator": "found_by",
        "recoverymethod": "recovery_method",
        "material": "material_id",
        "fabric": "material_text",
        "materialtext": "material_text",
        "technique": "technique",
        "decoration": "decoration",
        "inscription": "inscription",
        "length": "length_mm",
        "width": "width_mm",
        "height": "height_mm",
        "thickness": "thickness_mm",
        "diameter": "diameter_mm",
        "diam": "diameter_mm",
        "rimdiameter": "rim_diameter_mm",
        "rimdiam": "rim_diameter_mm",
        "weight": "weight_g",
        "period": "period_id",
        "periodtext": "period_text",
        "dating": "period_text",
        "datingmethod": "dating_method",
        "datefrom": "date_from",
        "dateto": "date_to",
        "condition": "condition",
        "conservation": "conservation_status",
        "conservationstatus": "conservation_status",
        "conservationnotes": "conservation_notes",
        "location": "current_location",
        "currentlocation": "current_location",
        "storage": "current_location",
        "box": "storage_box",
        "boxno": "storage_box",
        "ondisplay": "is_on_display",
        "notes": "research_notes",
        "remarks": "research_notes",
        "keywords": "keywords",
    },
}


def suggest_mapping(record_type: str, columns: list[str]) -> dict[str, str | None]:
    """Guess which field each column fills. Every guess is shown for approval.

    Three passes, most confident first: an exact synonym, an exact match on the
    field's own name or label, then a loose containment match. A column nothing
    matches maps to ``None``, which means "do not import this column" — the
    honest answer, and the safe one.
    """
    layout = forms.get_layout(record_type)
    if layout is None:
        raise SpreadsheetError(f"There is no import layout for {record_type!r}.")

    fields = forms.field_index(layout)
    by_name = {_slug(name): name for name in fields}
    by_label = {_slug(item.label): name for name, item in fields.items()}
    synonyms = SYNONYMS.get(record_type, {})

    mapping: dict[str, str | None] = {}
    taken: set[str] = set()

    for column in columns:
        key = _slug(column)
        guess = synonyms.get(key) or by_name.get(key) or by_label.get(key)

        if guess is None:
            # Loose match, but only when it is unambiguous: two candidates mean
            # we do not know, and a wrong guess a person waves through is worse
            # than no guess at all.
            candidates = {
                name
                for slug, name in {**by_name, **by_label}.items()
                if key and (key in slug or slug in key)
            }
            guess = candidates.pop() if len(candidates) == 1 else None

        # A field can only be filled once. The first column wins; the second is
        # left unmapped for a person to resolve.
        if guess is not None and guess in taken:
            guess = None
        if guess is not None:
            taken.add(guess)

        mapping[column] = guess

    return mapping


@dataclass
class ColumnReport:
    """What one column holds, and what it is proposed to fill."""

    column: str
    suggested_field: str | None
    field_label: str | None
    field_kind: str | None
    #: Values from the file, so a person can see what they are approving.
    samples: list[Any] = field(default_factory=list)
    #: How many of the sampled rows have anything in this column.
    filled: int = 0
    total: int = 0
    #: Set when the column is proposed for a field whose value list will not
    #: accept some of what is in it.
    warning: str | None = None


def describe_columns(
    record_type: str, sheet: Sheet, mapping: dict[str, str | None] | None = None
) -> list[ColumnReport]:
    """The verification screen's data: every column, what it holds, what it fills."""
    layout = forms.get_layout(record_type)
    if layout is None:
        raise SpreadsheetError(f"There is no import layout for {record_type!r}.")
    fields = forms.field_index(layout)

    resolved = mapping if mapping is not None else suggest_mapping(record_type, sheet.columns)

    reports = []
    for column in sheet.columns:
        target = resolved.get(column)
        spec = fields.get(target) if target else None

        values = [row.get(column) for row in sheet.rows]
        filled = [value for value in values if value not in (None, "")]

        reports.append(
            ColumnReport(
                column=column,
                suggested_field=target,
                field_label=spec.label if spec else None,
                field_kind=spec.kind if spec else None,
                samples=list(filled[:SAMPLE_ROWS]),
                filled=len(filled),
                total=len(values),
            )
        )
    return reports


def unmapped_required(record_type: str, mapping: dict[str, str | None]) -> list[str]:
    """Required fields no column fills, so the caller can say so before a run."""
    layout = forms.get_layout(record_type)
    if layout is None:
        return []
    filled = {value for value in mapping.values() if value}
    return [
        item.label
        for name, item in forms.field_index(layout).items()
        if item.required and not item.read_only and name not in filled
    ]
