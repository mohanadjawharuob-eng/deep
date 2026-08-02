"""Exporting a whole dataset as one Excel workbook.

A CSV is one table. But nobody's question is one table: "send me everything on
Tell el-Demo" means the site, its contexts, the stratigraphic relationships
between them, its finds, its photographs, its documents and its survey data —
and a folder of seven CSVs is not an answer, it is a filing job handed to
whoever asked.

So this builds one file with one sheet per kind of record, plus a cover sheet
that says what is in it and when it was made.

Four things make the difference between a workbook somebody can use and a dump
of the database.

**Identifiers are resolved to names.** A column holding
``a5ce6a8f-97b4-…`` is worthless in a spreadsheet. Every foreign key is
written as the thing it points at — the material's name, the period's name,
the context's number — and the raw id goes in a trailing column for anyone
re-importing it.

**Empty sheets are left out.** A site with no 3D models should not produce a
sheet called "3D models" with a header row and nothing under it; it makes the
reader hunt for content that was never there.

**The header row is frozen and the columns are sized.** A hundred-column
sheet where every column is eight characters wide is a sheet people give up
on in the first ten seconds.

**A cover sheet says what this is.** What was exported, by whom, when, on
what filters, and how many rows are in each sheet — so a file found on a
memory stick in two years can still explain itself.
"""

from __future__ import annotations

import io
import uuid
from collections.abc import Callable, Sequence
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal
from enum import Enum
from typing import Any

#: Sheet names Excel refuses, plus its 31-character limit, are handled in
#: :func:`_sheet_name`. These are the characters it will not accept at all.
_FORBIDDEN = set(r"[]:*?/\\")

#: Widest a column is allowed to get. Excel will happily make one 500
#: characters wide for a single long description, which pushes everything else
#: off the screen.
_MAX_WIDTH = 60
_MIN_WIDTH = 9


@dataclass(slots=True)
class Column:
    """One column: what it is called, and how to get it out of a record."""

    header: str
    #: Given the record, return the cell value. Defaults to the attribute of
    #: the same name, lower-cased and underscored.
    get: Callable[[Any], Any] | None = None
    #: Attribute to read when ``get`` is not given.
    attribute: str | None = None

    def value(self, record: Any) -> Any:
        if self.get is not None:
            return self.get(record)
        name = self.attribute or self.header.lower().replace(" ", "_")
        return getattr(record, name, None)


@dataclass(slots=True)
class Table:
    """One sheet: a title, its columns, and its rows."""

    title: str
    columns: list[Column]
    rows: Sequence[Any]
    #: Shown on the cover sheet under the row count. For saying "coordinates
    #: are blurred because this site is protected", which a reader has to know.
    note: str | None = None


@dataclass(slots=True)
class Workbook:
    """Everything that is going into the file."""

    #: What this export is of: "Tell el-Demo (TED)".
    subject: str
    #: What kind of thing that is: "Site", "Collection".
    kind: str
    tables: list[Table] = field(default_factory=list)
    #: Free-form lines for the cover sheet — the filters that were applied,
    #: the warning about restricted coordinates, anything a reader needs.
    notes: list[str] = field(default_factory=list)
    exported_by: str | None = None

    def add(self, table: Table) -> None:
        # Empty sheets are left out, not written blank. See the module
        # docstring: a heading with nothing under it sends the reader looking.
        if table.rows:
            self.tables.append(table)


def _cell(value: Any) -> Any:
    """Turn a column value into something openpyxl will write.

    Dates go in as real dates so Excel sorts and filters them; everything
    exotic becomes text, because a spreadsheet cell that says
    ``<Material object at 0x7f…>`` is worse than no cell at all.
    """
    # Enum first, and this ordering is load-bearing. Every enum in this schema
    # subclasses ``str``, so a plain ``isinstance(value, str)`` above would
    # match and hand openpyxl the member itself — which writes the cell as
    # "ConditionState.UNKNOWN" rather than "unknown". It looks like a fault in
    # the data rather than in the export, which is the worst way for a bug to
    # present in a file somebody sends to a funder.
    if isinstance(value, Enum):
        raw = value.value
        return raw.replace("_", " ") if isinstance(raw, str) else raw
    if value is None or isinstance(value, str | int | float | bool):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        # Excel cannot hold a timezone. Converting to UTC and dropping it is
        # honest as long as the header says so, which _write does.
        return value.astimezone(UTC).replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, date):
        return value
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, list | tuple | set):
        return ", ".join(str(_cell(item)) for item in value if item is not None)
    if isinstance(value, dict):
        return "; ".join(f"{key}: {_cell(item)}" for key, item in value.items() if item is not None)
    return str(value)


def _sheet_name(title: str, used: set[str]) -> str:
    """A name Excel will accept, and that is still recognisable.

    Excel allows 31 characters and refuses several punctuation marks. Silently
    truncating two long names to the same 31 characters makes the second sheet
    fail to be created at all, so collisions are numbered.
    """
    cleaned = "".join(" " if character in _FORBIDDEN else character for character in title).strip()
    cleaned = cleaned[:31] or "Sheet"

    candidate = cleaned
    suffix = 2
    while candidate.lower() in used:
        tail = f" {suffix}"
        candidate = cleaned[: 31 - len(tail)] + tail
        suffix += 1
    used.add(candidate.lower())
    return candidate


def build(workbook: Workbook) -> bytes:
    """Render a workbook to .xlsx bytes."""
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    book = openpyxl.Workbook()
    used: set[str] = set()

    # --- The cover ---------------------------------------------------------
    cover = book.active
    cover.title = _sheet_name("About this file", used)

    moment = datetime.now(UTC)
    lines: list[tuple[str, Any]] = [
        (f"{workbook.kind}", workbook.subject),
        ("Exported", moment.replace(tzinfo=None)),
        ("Exported by", workbook.exported_by or "unknown"),
        ("Made by", "Stratum"),
        ("", ""),
        ("Sheet", "Rows"),
    ]
    for table in workbook.tables:
        lines.append((table.title, len(table.rows)))
    if workbook.notes:
        lines.append(("", ""))
        lines.append(("Notes", ""))
        lines.extend(("", note) for note in workbook.notes)
    for table in workbook.tables:
        if table.note:
            lines.append(("", f"{table.title}: {table.note}"))

    for row_number, (label, value) in enumerate(lines, start=1):
        cover.cell(row=row_number, column=1, value=label).font = Font(bold=True)
        cover.cell(row=row_number, column=2, value=_cell(value))
    cover.column_dimensions["A"].width = 26
    cover.column_dimensions["B"].width = 80
    cover.cell(row=1, column=1).font = Font(bold=True, size=14)
    cover.cell(row=1, column=2).font = Font(bold=True, size=14)

    # --- One sheet per table ----------------------------------------------
    for table in workbook.tables:
        sheet = book.create_sheet(_sheet_name(table.title, used))

        for index, column in enumerate(table.columns, start=1):
            cell = sheet.cell(row=1, column=index, value=column.header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="top")

        widths = [len(column.header) for column in table.columns]
        for row_number, record in enumerate(table.rows, start=2):
            for index, column in enumerate(table.columns, start=1):
                value = _cell(column.value(record))
                sheet.cell(row=row_number, column=index, value=value)
                if value is not None:
                    widths[index - 1] = max(widths[index - 1], len(str(value)))

        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = min(
                max(width + 2, _MIN_WIDTH), _MAX_WIDTH
            )

        # The header stays put while somebody scrolls, and the filter arrows
        # are what makes a long sheet usable at all.
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(table.columns))}{len(table.rows) + 1}"

    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


def filename(kind: str, subject: str) -> str:
    """A filename that says what the file is, and that every system accepts."""
    import re

    stem = re.sub(r"[^\w.-]+", "-", f"{kind}-{subject}", flags=re.UNICODE).strip("-.")
    return f"{stem[:80] or 'export'}-{datetime.now(UTC):%Y%m%d}.xlsx"
