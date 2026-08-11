"""Exporting a whole dataset as one workbook.

The tests that matter are the two that make an export safe rather than
convenient: it must not become a side door around the permission model, and a
restricted site's coordinates must not walk out of the building in a
spreadsheet. Everything else is formatting.
"""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.models import Module, ModuleLevel, User, UserRole
from tests.conftest import auth_headers, make_user

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def open_book(content: bytes):
    import openpyxl

    return openpyxl.load_workbook(io.BytesIO(content))


def rows_of(book, sheet: str) -> list[dict]:
    """A sheet as dictionaries, keyed by its header row."""
    page = book[sheet]
    values = list(page.iter_rows(values_only=True))
    headers = [str(cell) if cell is not None else "" for cell in values[0]]
    return [dict(zip(headers, row, strict=False)) for row in values[1:]]


@pytest.fixture
def director(db: Session) -> User:
    return make_user(db, email="dir@example.org", username="dir", role=UserRole.ADMIN)


@pytest.fixture
def student(db: Session) -> User:
    return make_user(db, email="stu@example.org", username="stu", role=UserRole.STUDENT)


@pytest.fixture
def dig(client: TestClient, director: User) -> dict:
    """A project with a site, two contexts and two finds on it."""
    headers = auth_headers(client, "dir")

    project = client.post(
        "/api/v1/projects",
        json={"name": "Tell el-Demo", "code": "TED-2024"},
        headers=headers,
    ).json()

    site = client.post(
        "/api/v1/sites",
        json={
            "name": "Tell el-Demo",
            "code": "TED",
            "project_id": project["id"],
            "latitude": 34.7324,
            "longitude": 36.7137,
        },
        headers=headers,
    )
    assert site.status_code == 201, site.text
    site = site.json()

    contexts = []
    for number in ("1001", "1002"):
        created = client.post(
            "/api/v1/contexts",
            json={"context_number": number, "site_id": site["id"], "context_type": "layer"},
            headers=headers,
        )
        assert created.status_code == 201, created.text
        contexts.append(created.json())

    for number, name in (("TED-2024-0001", "Rim sherd"), ("TED-2024-0002", "Flint blade")):
        created = client.post(
            "/api/v1/artifacts",
            json={
                "inventory_number": number,
                "name": name,
                "site_id": site["id"],
                "context_id": contexts[0]["id"],
            },
            headers=headers,
        )
        assert created.status_code == 201, created.text

    return {"project": project, "site": site, "contexts": contexts}


class TestSiteWorkbook:
    def test_it_returns_a_real_spreadsheet(
        self, client: TestClient, director: User, dig: dict
    ) -> None:
        response = client.get(
            f"/api/v1/exports/sites/{dig['site']['id']}.xlsx",
            headers=auth_headers(client, "dir"),
        )
        assert response.status_code == 200, response.text
        assert response.headers["content-type"].startswith(XLSX)
        assert "attachment" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]

        book = open_book(response.content)
        assert "About this file" in book.sheetnames

    def test_one_sheet_per_kind_of_record(
        self, client: TestClient, director: User, dig: dict
    ) -> None:
        book = open_book(
            client.get(
                f"/api/v1/exports/sites/{dig['site']['id']}.xlsx",
                headers=auth_headers(client, "dir"),
            ).content
        )
        assert "Site" in book.sheetnames
        assert "Contexts" in book.sheetnames
        assert "Finds" in book.sheetnames

    def test_sheets_with_nothing_in_them_are_left_out(
        self, client: TestClient, director: User, dig: dict
    ) -> None:
        book = open_book(
            client.get(
                f"/api/v1/exports/sites/{dig['site']['id']}.xlsx",
                headers=auth_headers(client, "dir"),
            ).content
        )
        # Nothing uploaded, so no empty headings sending the reader hunting for
        # content that was never there.
        assert "Photographs" not in book.sheetnames
        assert "3D models" not in book.sheetnames

    def test_identifiers_are_written_as_names(
        self, client: TestClient, director: User, dig: dict
    ) -> None:
        book = open_book(
            client.get(
                f"/api/v1/exports/sites/{dig['site']['id']}.xlsx",
                headers=auth_headers(client, "dir"),
            ).content
        )
        finds = rows_of(book, "Finds")
        assert len(finds) == 2
        # The context reads as its number, not as a UUID nobody can use.
        assert finds[0]["Context"] == "1001"
        assert finds[0]["Site"] == "Tell el-Demo"
        # …and the raw identifier is still there for anybody re-importing.
        assert len(str(finds[0]["Identifier"])) == 36

    def test_the_cover_sheet_says_what_the_file_is(
        self, client: TestClient, director: User, dig: dict
    ) -> None:
        book = open_book(
            client.get(
                f"/api/v1/exports/sites/{dig['site']['id']}.xlsx",
                headers=auth_headers(client, "dir"),
            ).content
        )
        text = "\n".join(
            str(cell)
            for row in book["About this file"].iter_rows(values_only=True)
            for cell in row
            if cell is not None
        )
        assert "Tell el-Demo" in text
        assert "Stratum" in text
        # Who exported it: a file found on a memory stick has to explain itself.
        assert "dir" in text.lower()

    def test_enums_are_written_as_words(
        self, client: TestClient, director: User, dig: dict
    ) -> None:
        """Every enum in this schema subclasses ``str``.

        So a type check that asks "is this a string?" before "is this an
        enum?" matches, hands openpyxl the member itself, and the cell reads
        "ConditionState.UNKNOWN". In a file sent to a funder that looks like
        broken data rather than a broken export.
        """
        book = open_book(
            client.get(
                f"/api/v1/exports/sites/{dig['site']['id']}.xlsx",
                headers=auth_headers(client, "dir"),
            ).content
        )
        finds = rows_of(book, "Finds")
        assert finds[0]["Condition"] == "unknown"
        assert finds[0]["Review status"] == "approved"

        contexts = rows_of(book, "Contexts")
        # Underscores become spaces too: "single context" beats "single_context".
        assert "_" not in str(contexts[0]["Type"])

    def test_the_header_row_is_frozen(self, client: TestClient, director: User, dig: dict) -> None:
        book = open_book(
            client.get(
                f"/api/v1/exports/sites/{dig['site']['id']}.xlsx",
                headers=auth_headers(client, "dir"),
            ).content
        )
        assert book["Finds"].freeze_panes == "A2"


class TestExportsAreNotASideDoor:
    def test_a_site_you_cannot_see_is_not_exportable(
        self, client: TestClient, director: User, student: User, dig: dict
    ) -> None:
        """The same 404 the record itself would give.

        Anything else — a 403, say — confirms the site exists, which for a
        protected site is itself the information being protected.
        """
        client.patch(
            f"/api/v1/sites/{dig['site']['id']}",
            json={"is_public": False},
            headers=auth_headers(client, "dir"),
        )
        response = client.get(
            f"/api/v1/exports/sites/{dig['site']['id']}.xlsx",
            headers=auth_headers(client, "stu"),
        )
        assert response.status_code == 404

    def test_it_needs_an_account(self, client: TestClient, director: User, dig: dict) -> None:
        assert client.get(f"/api/v1/exports/sites/{dig['site']['id']}.xlsx").status_code == 401

    def test_a_restricted_location_stays_restricted(
        self, client: TestClient, director: User, db: Session, dig: dict
    ) -> None:
        """Blurring on the map and printing the truth in a file protects nothing."""
        client.patch(
            f"/api/v1/sites/{dig['site']['id']}",
            json={"location_restricted": True, "is_public": True},
            headers=auth_headers(client, "dir"),
        )

        # Somebody who may export, but is not on this project and so cannot
        # edit the site. Exporting needs contributor level; a plain viewer
        # cannot export at all, which is a separate rule and not what this
        # test is about.
        make_user(
            db,
            email="visitor2@example.org",
            username="visitor2",
            role=UserRole.VISITOR,
            modules={Module.ARCHAEOLOGY: ModuleLevel.CONTRIBUTOR},
            grant_defaults=False,
        )

        response = client.get(
            f"/api/v1/exports/sites/{dig['site']['id']}.xlsx",
            headers=auth_headers(client, "visitor2"),
        )
        assert response.status_code == 200, response.text
        book = open_book(response.content)
        site_rows = rows_of(book, "Site")
        assert site_rows[0]["Latitude"] is None
        assert site_rows[0]["Longitude"] is None

        # And the cover says so, because a silent blank reads as missing data.
        text = "\n".join(
            str(cell)
            for row in book["About this file"].iter_rows(values_only=True)
            for cell in row
            if cell is not None
        )
        assert "restricted" in text.lower()

    def test_an_editor_still_gets_the_real_coordinates(
        self, client: TestClient, director: User, dig: dict
    ) -> None:
        client.patch(
            f"/api/v1/sites/{dig['site']['id']}",
            json={"location_restricted": True},
            headers=auth_headers(client, "dir"),
        )
        book = open_book(
            client.get(
                f"/api/v1/exports/sites/{dig['site']['id']}.xlsx",
                headers=auth_headers(client, "dir"),
            ).content
        )
        assert rows_of(book, "Site")[0]["Latitude"] == pytest.approx(34.7324)


class TestProjectWorkbook:
    def test_it_gathers_every_site_into_one_sheet_per_kind(
        self, client: TestClient, director: User, dig: dict
    ) -> None:
        response = client.get(
            f"/api/v1/exports/projects/{dig['project']['id']}.xlsx",
            headers=auth_headers(client, "dir"),
        )
        assert response.status_code == 200, response.text

        book = open_book(response.content)
        assert "Project" in book.sheetnames
        assert "Sites" in book.sheetnames
        # Not "Tell el-Demo finds", "Tell B finds", … — forty sheets nobody can
        # filter across.
        assert "Finds" in book.sheetnames
        assert len(rows_of(book, "Finds")) == 2

    def test_the_finds_name_their_site(self, client: TestClient, director: User, dig: dict) -> None:
        book = open_book(
            client.get(
                f"/api/v1/exports/projects/{dig['project']['id']}.xlsx",
                headers=auth_headers(client, "dir"),
            ).content
        )
        assert {row["Site"] for row in rows_of(book, "Finds")} == {"Tell el-Demo"}
